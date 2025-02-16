from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border
from datetime import datetime, timedelta
from calendar import monthrange
import os
import logging
from utils.utils import PUBLIC_HOLIDAYS, load_user_details  # Import function instead of USER_DETAILS
from styles import (  # Import styles from styles.py
    thin_border, white_fill, yellow_fill, light_green_fill, lighter_green_fill, light_yellow_fill, light_blue_fill,
    light_red_fill, bold_font, red_font, black_font, center_alignment, right_alignment)

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

def generate_timesheet_excel(user_id, month, year, leave_details):
    USER_DETAILS = load_user_details()
    user_details = USER_DETAILS.get(user_id)
    if not user_details:
        raise ValueError(f"User with ID {user_id} not found.")

    ns_leave_present = any(leave_type == "NS Leave" for _, _, leave_type in leave_details)


    name = user_details["name"]
    skill_level = user_details["skill_level"]
    role_specialization = user_details["role_specialization"]
    group_specialization = user_details["group_specialization"]
    contractor = user_details["contractor"]
    po_ref = user_details["po_ref"]
    po_date = user_details["po_date"]
    description = user_details["description"]
    reporting_officer = user_details["reporting_officer"]
    # Fetch timesheet preference (Default to 1.0 if not set)
    timesheet_preference = float(user_details.get("timesheet_preference", 1.0))

    # **Check if NS Leave exists in the given month**
   # ns_leave_present = any(leave_type == "NS Leave" for _, leave_type in leave_details)

    # File Setup
    month_name = datetime(year, month, 1).strftime("%B")
    filename = f"{month_name}_{year}_Timesheet_{name.replace(' ', '_')}.xlsx"
    output_dir = "generated_timesheets"
    os.makedirs(output_dir, exist_ok=True)
    output_file = os.path.join(output_dir, filename)

    # Workbook & Worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = f"{month_name} {year} Timesheet"


    ws.merge_cells("B2:D2")

    # Set row height for Description row to ensure proper spacing
    ws.row_dimensions[8].height = 25  # Adjust row height for better text display

    # Set a fixed width for the merged column
    ws.column_dimensions["B"].width = 25  # Adjust width slightly larger than text
    ws.column_dimensions["C"].width = 10  # Prevent unwanted expansion
    ws.column_dimensions["D"].width = 10

    # Enable wrap text
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells("B3:D3")  # Merge PO Ref value
    ws.merge_cells("B4:D4")  # Merge PO Date value
    ws.merge_cells("G2:H2")  # Merge Month/Year value
    ws.merge_cells("G3:H3")  # Merge Contractor value

    ws["A2"], ws["B2"] = "Description", description
    ws["A3"], ws["B3"] = "PO Ref", po_ref
    ws["A4"], ws["B4"] = "PO Date", po_date
    ws["F2"], ws["G2"] = "Month/Year", f"{month_name} - {year}"
    ws["F3"], ws["G3"] = "Contractor", contractor

    # Apply Borders for Header Sections
    for row in range(2, 5):  # A-D Borders for Description, PO Ref, PO Date
        for col in ["A", "B", "C", "D"]:
            ws[f"{col}{row}"].border = thin_border

    for row in range(2, 4):  # F-H Borders for Month/Year and Contractor
        for col in ["F", "G", "H"]:
            ws[f"{col}{row}"].border = thin_border

    # Apply Borders for Month/Year (Fully Inside a Box)
    for col in ["F", "G", "H"]:
        ws[f"{col}2"].border = thin_border

    # Apply Yellow Fill to Values Only (B-D, G-H)
    for row in range(2, 5):
        for col in ["B", "C", "D"]:
            ws[f"{col}{row}"].fill = yellow_fill  # Apply to Description, PO Ref, PO Date

    # Apply Yellow Fill ONLY to Month/Year (G2:H2)
    for col in ["G", "H"]:
        ws[f"{col}2"].fill = yellow_fill

        # Ensure Contractor (G3:H3) is NOT Yellow
    for col in ["G", "H"]:
        ws[f"{col}3"].fill = PatternFill(fill_type=None)  # Remove yellow fill

    # Ensure Column E is empty (No Borders, No Fill)
    for row in range(2, 5):
        ws[f"E{row}"].border = Border()
        ws[f"E{row}"].fill = PatternFill(fill_type=None)

    # **Apply Bottom Alignment for Values (B-D)**
    for row in range(2, 5):
        for col in ["B", "C", "D"]:
            ws[f"{col}{row}"].alignment = Alignment(horizontal="center", vertical="bottom")

    # **Keep Column A Left-Aligned**
    for row in range(2, 5):
        ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="bottom")

    # **Apply Center Alignment for Skill Level, Contractor, and Month/Year Values**
    for row in [2, 3, 6]:  # Month/Year, Contractor, Skill Level
        for col in ["G", "H"]:
            ws[f"{col}{row}"].alignment = Alignment(horizontal="center", vertical="bottom")

    # **Wrap Text for Description (B2:D2)**
    # Set fixed column width for Description (B2:D2) and enable wrap text
    description_columns = ["B", "C", "D"]
    for col in description_columns:
        ws.column_dimensions[col].width = 15  # Adjust width (set slightly larger than text)
        ws[f"{col}2"].alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)

    # Merge and Format User Details
    ws.merge_cells("B6:D6")  # Merge Name
    ws.merge_cells("B7:D7")  # Merge Role Specialization
    ws.merge_cells("B8:D8")  # Merge Group Specialization
    ws.merge_cells("G6:H6")  # Merge Skill Level

    ws["A6"], ws["B6"] = "Name", name
    ws["A7"], ws["B7"] = "Role Specialization", role_specialization
    ws["A8"], ws["B8"] = "Group/Specialization", group_specialization
    ws["F6"], ws["G6"] = "Skill Level", skill_level

    # Apply Borders for User Details
    for row in range(6, 9):  # A-D Borders for Name, Role Specialization, Group
        for col in ["A", "B", "C", "D"]:
            ws[f"{col}{row}"].border = thin_border

    for row in [6]:  # F-H Borders for Skill Level
        for col in ["F", "G", "H"]:
            ws[f"{col}{row}"].border = thin_border

    # Apply Yellow Fill to Values Only (B-D, G-H)
    for row in range(6, 9):
        for col in ["B", "C", "D"]:
            ws[f"{col}{row}"].fill = yellow_fill  # Name, Role, Group

    for row in [6]:
        for col in ["G", "H"]:
            ws[f"{col}{row}"].fill = yellow_fill  # Skill Level

    # Ensure Column E is empty (No Borders, No Fill)
    for row in range(6, 9):
        ws[f"E{row}"].border = Border()
        ws[f"E{row}"].fill = PatternFill(fill_type=None)

    # **Left Align Name, Role Specialization, Group Specialization Values**
    for row in range(6, 9):
        ws[f"B{row}"].alignment = Alignment(horizontal="left", vertical="bottom")  # Left align

    # # Table Headers
    # headers = ["SN", "Date", "At Work", "Public Holiday", "Sick Leave", "Childcare Leave", "Annual Leave", "Remarks"]
    # **Table Headers**
    headers = ["SN", "Date", "At Work", "Public Holiday", "Sick Leave", "Childcare Leave", "Annual Leave"]
    if ns_leave_present:
        #headers.append("NS Leave")  # Add NS Leave column only if applicable
        headers.append("National Service Leave")  # Rename "NS Leave" if required
    headers.append("Remarks")  # Add Remarks at the end

    header_fills = [
        white_fill, white_fill, light_green_fill, light_yellow_fill, lighter_green_fill, white_fill, light_blue_fill
    ]
    if ns_leave_present:
        header_fills.append(light_red_fill)  # Color for NS Leave
    header_fills.append(white_fill)

    # **Create Table Headers Dynamically**
    for col_num, (header, fill) in enumerate(zip(headers, header_fills), 1):
        cell = ws.cell(row=10, column=col_num, value=header)

        # Apply font and styling dynamically
        is_sn_column = col_num == 1  # SN column
        cell.font = Font(name="Arial", size=12, bold=not is_sn_column, color="000000")  # SN should be non-bold
        cell.alignment = Alignment(horizontal="center", vertical="center")  # Center alignment
        cell.border = thin_border  # Apply thin border
        cell.fill = fill  # Apply respective color

    # Adjust column width dynamically based on NS Leave presence
    if ns_leave_present:
        ws.column_dimensions["I"].width = 12  # Ensure NS Leave column has proper width
        remarks_col = "J"  # Remarks moves to J
    else:
        remarks_col = "I"  # Remarks remains in I

    # Ensure SN column (A10) matches the Remarks column formatting (non-bold, middle-aligned)
    ws["A10"].font = Font(name="Arial", size=12, bold=False, color="000000")  # Make SN non-bold
    ws["A10"].alignment = Alignment(horizontal="center", vertical="center")  # Middle-aligned like Remarks
    ws["A10"].border = thin_border  # Keep border styling

    # # Ensure "Remarks" header (Column 8) is formatted the same way
    # ws["H10"].font = Font(name="Arial", size=12, bold=False, color="000000")  # Match font
    # Determine the correct column letter for Remarks based on NS Leave presence
    remarks_column_letter = "I" if ns_leave_present else "H"
    # Ensure "Remarks" header is formatted the same way
    ws[f"{remarks_column_letter}10"].font = Font(name="Arial", size=12, bold=False, color="000000")  # Match font

    expanded_leave_details = []

    try:
        for leave_entry in leave_details:
            if isinstance(leave_entry, tuple) and len(leave_entry) == 3:
                start_date, end_date, leave_type = leave_entry
                logging.info(f"Expanding leave range: {start_date} to {end_date} ({leave_type})")

                start_date = datetime.strptime(start_date, "%d-%B").replace(year=year)
                end_date = datetime.strptime(end_date, "%d-%B").replace(year=year)

                while start_date <= end_date:
                    expanded_leave_details.append((start_date.strftime("%Y-%m-%d"), leave_type))  # NS Leave Included
                    start_date += timedelta(days=1)


            elif isinstance(leave_entry, tuple) and len(leave_entry) == 2:
                # Direct (date, leave_type) entry
                date_str, leave_type = leave_entry
                try:
                    formatted_date = datetime.strptime(date_str, "%d-%B").replace(year=year).strftime("%Y-%m-%d")
                    expanded_leave_details.append((formatted_date, leave_type))
                except ValueError:
                    logging.error(f"Invalid date format in leave entry: {leave_entry}")

            else:
                logging.error(f"Unexpected leave format: {leave_entry}")
                continue  # Skip invalid entries

        logging.info(f"Final expanded leave details: {expanded_leave_details}")

    except Exception as e:
        logging.error(f"Error processing leave details: {e}")

    # **Data Rows**
    current_row = 11

    sn_counter = 1  # Start SN from 1

    _, days_in_month = monthrange(year, month)
    totals = {"At Work": 0.0, "Public Holiday": 0.0, "Sick Leave": 0.0, "Childcare Leave": 0.0, "Annual Leave": 0.0}
    if ns_leave_present:
        #totals["NS Leave"] = 0.0  # Initialize NS Leave Total
        totals["National Service Leave"] = 0.0  # Rename "NS Leave" to "National Service Leave"

    for day in range(1, days_in_month + 1):
        date_obj = datetime(year, month, day)
        formatted_date = date_obj.strftime("%d-%B-%Y")  # Display format
        public_holiday_check = date_obj.strftime("%Y-%m-%d")  # Match keys in PUBLIC_HOLIDAYS
        weekday = date_obj.weekday()

        # Set "At Work" value dynamically based on timesheet preference
        at_work = timesheet_preference if weekday not in [5, 6] else 0.0
        # public_holiday, sick_leave, childcare_leave, annual_leave = 0.0, 0.0, 0.0, 0.0  # Default to user preference, except weekends
        public_holiday, sick_leave, childcare_leave, annual_leave, ns_leave = 0.0, 0.0, 0.0, 0.0, 0.0
        remark = "-"  # Default empty remark
        # **Handle Weekends**
        if weekday == 5:
            at_work = 0.0
            remark = "Saturday"

        if weekday in [5, 6]:
            at_work = 0.0
            remark = "Weekend"

        if public_holiday_check in PUBLIC_HOLIDAYS:
            at_work = 0.0
            public_holiday = 1.0
            remark = PUBLIC_HOLIDAYS[public_holiday_check]

        for leave_date, leave_type in expanded_leave_details:
            logger.debug(f"Processing Leave Date: {leave_date}, Type: {leave_type}")
            if leave_date == date_obj.strftime("%Y-%m-%d"):
                if leave_type == "Sick Leave":
                    # Sick Leave should NOT apply on weekends or public holidays
                    if weekday not in [5, 6] and public_holiday_check not in PUBLIC_HOLIDAYS:
                        sick_leave = 1.0
                        at_work = 0.0
                elif leave_type == "Childcare Leave":
                    # Childcare Leave should NOT apply on weekends or public holidays
                    if weekday not in [5, 6] and public_holiday_check not in PUBLIC_HOLIDAYS:
                        childcare_leave = 1.0
                        at_work = 0.0
                elif leave_type == "Annual Leave":
                    # Annual Leave should NOT apply on weekends or public holidays
                    if weekday not in [5, 6] and public_holiday_check not in PUBLIC_HOLIDAYS:
                        annual_leave = 1.0
                        at_work = 0.0
                elif leave_type == "NS Leave":
                    if weekday not in [5, 6] and public_holiday_check not in PUBLIC_HOLIDAYS:
                        ns_leave = 1.0
                        at_work = 0.0  # No work on NS Leave
                elif leave_type == "Weekend Efforts":
                    # Only update at_work if it's a Saturday, Sunday, or Public Holiday
                    if weekday in [5, 6] or public_holiday_check in PUBLIC_HOLIDAYS:
                        at_work = 8.0 if timesheet_preference == 8.5 else 1.0
                elif leave_type == "Public Holiday Efforts":
                    # Only update at_work if it's a Public Holiday
                    if public_holiday_check in PUBLIC_HOLIDAYS:
                        at_work = 8.0 if timesheet_preference == 8.5 else 1.0
                elif leave_type == "Half Day":
                    # Half Day Handling:
                    # - If timesheet preference is 8.5:
                    #   - Monday to Thursday: Half day = 4.5 hours
                    #   - Friday: Half day = 4.0 hours
                    # - If timesheet preference is 1.0:
                    #   - Monday to Friday: Half day = 0.5 hours

                    if timesheet_preference == 8.5:
                        at_work = 4.5 if weekday not in [4] else 4.0  # Friday (4) gets 4.0 hours
                    else:
                        at_work = 0.5  # All weekdays get 0.5 hours

        totals["At Work"] += at_work if isinstance(at_work, float) else 0.0
        totals["Public Holiday"] += public_holiday
        totals["Sick Leave"] += sick_leave
        totals["Childcare Leave"] += childcare_leave
        totals["Annual Leave"] += annual_leave
        if ns_leave_present:
            #totals["NS Leave"] += ns_leave
            totals["National Service Leave"] += ns_leave  # Change "NS Leave" to "National Service Leave"

        # row_data = [current_row - 1, formatted_date, at_work, public_holiday, sick_leave, childcare_leave, annual_leave]
        # if ns_leave_present:
        #     row_data.append(ns_leave)  # Add NS Leave column data
        # row_data.append(remark)  # Always add Remarks
        # Replace 0.0 with an empty string to keep cells blank instead of showing 0.0
        row_data = [
            #current_row - 1,
            sn_counter,  # Now SN starts at 1, 2, 3...
            formatted_date,
            "" if at_work == 0.0 else at_work,
          #  "" if public_holiday == 0.0 else public_holiday,
            "-" if public_holiday == 0.0 else public_holiday,
            "" if sick_leave == 0.0 else sick_leave,
            "" if childcare_leave == 0.0 else childcare_leave,
            "" if annual_leave == 0.0 else annual_leave,
        ]

        if ns_leave_present:
            row_data.append("" if ns_leave == 0.0 else ns_leave)  # Handle NS Leave blank cells

        row_data.append(remark)  # Always add Remarks

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col_num, value=value)
            cell.alignment = center_alignment
            cell.border = thin_border
            # # Highlight Leave & Public Holiday Cells
            # if col_num in [3, 4, 5, 6, 7]:  # At Work, Public Holiday, Sick Leave, Childcare Leave, Annual Leave
            #     cell.fill = yellow_fill if value not in ["", "-"] else white_fill

            # Determine the correct column index for NS Leave based on its presence
            ns_leave_column_index = 8 if ns_leave_present else None

            # Highlight Leave & Public Holiday Cells
            leave_columns = [3, 4, 5, 6, 7]  # At Work, Public Holiday, Sick Leave, Childcare Leave, Annual Leave
            if ns_leave_present:
                leave_columns.append(ns_leave_column_index)  # Add NS Leave column if present

            if col_num in leave_columns:
                cell.fill = yellow_fill if value not in ["", "-"] else white_fill

            # Highlight Remarks for Public Holidays & Leaves | remove redudant  code
            # if col_num == 8 and remark not in ["-", ""]:
            #     cell.fill = light_red_fill
            # #   cell.font = bold_font
            # Determine the correct column index for Remarks based on NS Leave presence
            remarks_column_index = 9 if ns_leave_present else 8

            # Highlight Remarks for Public Holidays & Leaves
            if col_num == remarks_column_index and remark not in ["-", ""]:
                cell.fill = light_red_fill

            leave_columns = [3, 4, 5, 6, 7]  # At Work, Public Holiday, Sick Leave, Childcare Leave, Annual Leave
            if ns_leave_present:
                leave_columns.append(8)  # NS Leave column (H)

            for col_num in leave_columns:
                ws.cell(row=current_row, column=col_num).number_format = "0.0"  # Ensure 1 decimal place

        # Remove yellow for PH column and add yellow_fill to Date column
        # # Apply Yellow Fill to "Date" Column (B) but NOT to Public Holiday (D)
        # for row in range(11, 11 + days_in_month):  # Assuming row 11 is the first data row
        #     ws[f"B{row}"].fill = yellow_fill  # Apply Yellow Fill to Date Column
        #
        # # Apply Yellow Fill for "At Work", "Sick Leave", "Childcare Leave", and "Annual Leave" Columns (C, E, F, G)
        # for row in range(11, 11 + days_in_month):
        #     for col_num in [3, 5, 6, 7]:  # C=At Work, E=Sick Leave, F=Childcare Leave, G=Annual Leave
        #         cell = ws.cell(row=row, column=col_num)
        #         cell.fill = yellow_fill
        #
        # # Ensure "Public Holiday" (D) is NOT Yellow
        # for row in range(11, 11 + days_in_month):
        #     ws[f"D{row}"].fill = PatternFill(fill_type=None)  # Remove yellow fill from Public Holiday
        #
        # # Apply right aligned for At Work, Public Holiday, Sick Leave, Childcare Leave, and Annual Leave up to row 31
        # for row in range(11, 11 + days_in_month):  # Assuming row 11 is the first data row, row 41 is the last (31st day)
        #     for col_num in [3, 4, 5, 6,
        #                     7]:  # Columns: At Work (C), Sick Leave (E), Childcare Leave (F), Annual Leave (G)
        #         cell = ws.cell(row=row, column=col_num)
        #         cell.alignment = right_alignment
        #
        # for row in range(11, 53):  # Adjusting for row range from 11 to 42 (inclusive)
        #     cell = ws.cell(row=row, column=8)  # Column 8 is "Remarks"
        #     if cell.value not in ["-", ""]:  # Apply styles only to meaningful values
        #         cell.font = red_font
        #         cell.alignment = right_alignment  # Apply right alignment
        #     else:  # If the value is "-", keep it black
        #         cell.font = black_font
        #         cell.alignment = right_alignment
        # Apply Yellow Fill to "Date" Column (B) but NOT to Public Holiday (D)
        for row in range(11, 11 + days_in_month):  # Assuming row 11 is the first data row
            ws[f"B{row}"].fill = yellow_fill  # Apply Yellow Fill to Date Column

        # Determine the correct column indexes dynamically
        at_work_col, public_holiday_col, sick_leave_col, childcare_leave_col, annual_leave_col = 3, 4, 5, 6, 7
        ns_leave_col = 8 if ns_leave_present else None  # NS Leave column is 8 if present, else None
        remarks_col = 9 if ns_leave_present else 8  # Remarks column shifts to 9 if NS Leave exists

        # Apply Yellow Fill for "At Work", "Sick Leave", "Childcare Leave", Annual Leave, and NS Leave (if applicable)
        leave_columns = [at_work_col, sick_leave_col, childcare_leave_col, annual_leave_col]
        if ns_leave_present:
            leave_columns.append(ns_leave_col)  # Include NS Leave in yellow fill columns

        for row in range(11, 11 + days_in_month):
            for col_num in leave_columns:  # Apply Yellow Fill to all leave-related columns
                cell = ws.cell(row=row, column=col_num)
                cell.fill = yellow_fill

        # Ensure "Public Holiday" (D) is NOT Yellow
        for row in range(11, 11 + days_in_month):
            ws[f"D{row}"].fill = PatternFill(fill_type=None)  # Remove yellow fill from Public Holiday column

        # Apply right alignment for At Work, Public Holiday, Sick Leave, Childcare Leave, Annual Leave, and NS Leave (if applicable)
        align_columns = [at_work_col, public_holiday_col, sick_leave_col, childcare_leave_col, annual_leave_col]
        if ns_leave_present:
            align_columns.append(ns_leave_col)  # Add NS Leave if present

        for row in range(11, 11 + days_in_month):
            for col_num in align_columns:
                cell = ws.cell(row=row, column=col_num)
                cell.alignment = right_alignment  # Apply right alignment

        # Apply formatting to Remarks column dynamically
        for row in range(11, 53):  # Adjusting for row range from 11 to 42 (inclusive)
            cell = ws.cell(row=row, column=remarks_col)  # Dynamic Remarks column
            if cell.value not in ["-", ""]:  # Apply styles only to meaningful values
                cell.font = red_font
                cell.alignment = right_alignment  # Apply right alignment
            else:  # If the value is "-", keep it black
                cell.font = black_font
                cell.alignment = right_alignment

        sn_counter += 1
        current_row += 1

    current_row += 2
    # fixing error text
    # Apply "Total" Label
    # ws[f"A{current_row}"] = "Total"
    # ws[f"A{current_row}"].font = Font(name="Arial", size=12, bold=True, color="000000")  # Only "Total" is bold
    # ws[f"A{current_row}"].alignment = center_alignment
    total_cell = ws[f"A{current_row}"]
    total_cell.value = "Total"
    total_cell.font = Font(name="Arial", size=12, bold=True,
                           color="000000")  # Ensure bold is applied after value is set
    total_cell.alignment = center_alignment
    total_cell.border = thin_border

    # Apply Borders and Fix Number Format for Total Row (Columns C to H)
    # Determine column indexes dynamically
    remarks_column = 9 if ns_leave_present else 8  # "I" if NS Leave exists, else "H"
    ns_leave_column = 8 if ns_leave_present else None  # "H" if NS Leave exists, else None

    for col_num, key in enumerate(totals.keys(), 3):  # Starts from column C (At Work)
        total_value = totals[key]
        display_total = "-" if total_value == 0.0 else total_value  # Keep numbers as numbers
        cell = ws.cell(row=current_row, column=col_num, value=display_total)

        # Ensure total values remain **normal** (not bold)
        cell.font = Font(name="Arial", size=12, bold=False, color="000000")
        cell.alignment = right_alignment  # Apply right alignment
        cell.border = thin_border  # Apply border to each cell
        cell.number_format = "0.0"  # Ensure it's stored as a number (1 decimal place)

    # Apply border to the Remarks column dynamically
    ws[f"A{current_row}"].border = thin_border
    ws[f"B{current_row}"].border = thin_border
    ws[f"{chr(64 + remarks_column)}{current_row}"].border = thin_border  # Apply border dynamically based on Remarks column

    # **Signature Section**
    current_date = datetime.now().strftime("%d - %b - %Y")  # Ensure proper formatting before writing to Excel

    # Merge Officer Fields Across B, C, D
    ws.merge_cells(f"B{current_row + 2}:D{current_row + 2}")  # Officer Name
    ws.merge_cells(f"B{current_row + 3}:D{current_row + 3}")  # Officer Signature
    ws.merge_cells(f"B{current_row + 4}:D{current_row + 4}")  # Officer Date

    # Merge Reporting Officer Fields Across B, C, D
    ws.merge_cells(f"B{current_row + 6}:D{current_row + 6}")  # Reporting Officer Name
    ws.merge_cells(f"B{current_row + 7}:D{current_row + 7}")  # Empty Signature Field
    ws.merge_cells(f"B{current_row + 8}:D{current_row + 8}")  # Empty Date Field

    # Assign Values
    ws[f"A{current_row + 2}"], ws[f"B{current_row + 2}"] = "Officer", name
    ws[f"A{current_row + 3}"], ws[f"B{current_row + 3}"] = "Signature", name
    ws[f"A{current_row + 4}"], ws[f"B{current_row + 4}"] = "Date", current_date  # Date formatted

    ws[f"A{current_row + 6}"], ws[f"B{current_row + 6}"] = "Reporting Officer", reporting_officer
    ws[f"A{current_row + 7}"], ws[f"B{current_row + 7}"] = "Signature", ""  # Leave Empty for Manager
    ws[f"A{current_row + 8}"], ws[f"B{current_row + 8}"] = "Date", ""  # Leave Empty for Manager

    # **Apply Date Formatting**
    ws[f"B{current_row + 4}"].number_format = "DD - MMM - YYYY"  # Ensure date appears correctly
    ws[f"B{current_row + 8}"].number_format = "DD - MMM - YYYY"  # Format empty date field

    # Apply Borders to A-D
    for row in range(current_row + 2, current_row + 9):  # Covers Officer + Reporting Officer sections
        for col in ["A", "B", "C", "D"]:
            ws[f"{col}{row}"].border = thin_border

    # **Apply Bottom Alignment**
    for row in range(current_row + 2, current_row + 9):
        for col in ["A", "B", "C", "D"]:
            ws[f"{col}{row}"].alignment = Alignment(horizontal="center", vertical="bottom")

    # Keep A Column Left-Aligned
    for row in range(current_row + 2, current_row + 9):
        ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="bottom")

    # Determine the correct column index for Remarks based on NS Leave presence
    remarks_column_index = 9 if ns_leave_present else 8  # 9 = "I", 8 = "H"

    arial_font = Font(name="Arial", size=12)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            # Apply Arial font to all columns except the dynamically determined Remarks column
            if cell.column != remarks_column_index and cell.row != current_row:
                cell.font = arial_font  # Apply Arial 12 font

    # Determine Remarks column index based on NS Leave presence
    remarks_column = 9 if ns_leave_present else 8

    for row in range(11, 11 + days_in_month):
        cell = ws.cell(row=row, column=remarks_column)  # Remarks column
        date_cell = ws.cell(row=row, column=2)  # Date column to determine the weekday

        if date_cell.value:
            try:
                date_obj = datetime.strptime(date_cell.value, "%d-%B-%Y")  # Convert date to object
                weekday = date_obj.weekday()  # Get weekday (0 = Monday, 6 = Sunday)
            except ValueError:
                weekday = None  # In case date format is incorrect

        if cell.value and cell.value not in ["-", ""]:  # Apply styles only to meaningful values
            cell_value = str(cell.value).strip().lower()

            # Check if the remark is a public holiday or weekend (Saturday/Sunday)
            if any(holiday.lower() in cell_value for holiday in PUBLIC_HOLIDAYS.values()) or weekday in [5, 6]:
                cell.font = Font(name="Arial", size=12, color="FF0000", bold=False)  # Keep red font for PH & weekends
            else:
                cell.font = Font(name="Arial", size=12, color="000000", bold=False)  # Apply Arial 12 for other remarks
        else:  # If the value is "-", keep it black and in Arial
            cell.font = Font(name="Arial", size=12, color="000000", bold=False)

    ws.column_dimensions["B"].width = 25  # Keep Description column at a reasonable width

    # Auto-expand C and D (At Work, Public Holiday)
    for col in ["C", "D"]:
        max_length = max(len(str(cell.value or "")) for cell in ws[col])
        ws.column_dimensions[col].width = max(max_length + 2, 10)  # Ensure at least 10 width

    # Set other columns dynamically based on their content
    for col in ws.columns:
        col_letter = col[0].column_letter
        if col_letter not in ["B", "C", "D"]:  # Skip setting widths for manually controlled columns
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col_letter].width = min(max_length + 2, 20)  # Prevent over-expansion

    # Adjust the Remarks column (H) width based on the largest text in any column (A to H)
    ws.column_dimensions["H"].width = max(
        len(str(cell.value)) for row in ws.iter_rows(min_row=11, max_row=11 + days_in_month, min_col=1, max_col=8) for
        cell in row) + 2

    # **Save File**
    wb.save(output_file)
    print(f"Timesheet saved -> {output_file}")
    return output_file
