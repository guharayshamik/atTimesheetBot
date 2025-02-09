from openpyxl.styles import Alignment, PatternFill, Font, Border, Side

# Borders
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))

# Fills
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
light_green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light Green (At Work)
lighter_green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Lighter Green (Sick Leave)
light_yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Light Yellow (Public Holiday)
light_blue_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")  # Light Blue (Annual Leave)
light_red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light Red for remarks
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White (Default)

# Fonts
bold_font = Font(bold=True)  # Bold font
red_font = Font(color="FF0000")  # Red font for important remarks
black_font = Font(color="000000")  # Default black font

# Alignments
center_alignment = Alignment(horizontal="center", vertical="center")  # Center alignment
right_alignment = Alignment(horizontal="right", vertical="center")  # Right alignment
